"""
Bond-Defense 游戏配置表生成脚本
运行方式: python create_tables.py
会在 data/ 目录下生成英雄、羁绊、波次、敌人、商店配置的 xlsx 文件
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 颜色定义
COLOR_TYPE_ROW = "4472C4"      # 蓝色 - 类型行背景
COLOR_COMMENT_ROW = "70AD47"   # 绿色 - 注释行背景
COLOR_FIELD_ROW = "FFC000"     # 橙色 - 字段名行背景
COLOR_DATA_ODD = "FFFFFF"      # 白色 - 奇数数据行
COLOR_DATA_EVEN = "F2F2F2"     # 浅灰 - 偶数数据行
COLOR_HEADER_FONT = "FFFFFF"   # 白色字体（用于前三行）

def apply_header_style(ws, row, bg_color):
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    font = Font(name="微软雅黑", bold=True, color=COLOR_HEADER_FONT, size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = align

def apply_data_style(ws, row, even=False):
    bg = COLOR_DATA_EVEN if even else COLOR_DATA_ODD
    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    font = Font(name="微软雅黑", size=10)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = align

def set_col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_header(ws):
    ws.freeze_panes = "A4"  # 冻结前3行（类型/注释/字段名）

# ============================================================
# 表 1: 英雄配置表 (hero.xlsx)
# ============================================================
def create_hero_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "hero"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # 第1行：类型
    types = ["int",    "string",  "string",  "int",      "int",    "float",   "float",    "int",   "int",    "array_str",       "string"]
    # 第2行：注释
    comments = ["编号", "英雄名称", "稀有度",   "售价(金币)", "攻击力",  "攻速(次/秒)", "攻击范围(px)", "生命值", "标签数量",  "标签列表(|分隔)",   "颜色(RRGGBB)"]
    # 第3行：字段名
    fields = ["id",   "name",    "rarity",  "cost",     "attack", "attack_speed", "range", "hp",    "*tag_count", "tags",            "color"]
    # 数据行
    data = [
        [1, "人类弓手",  "common",    1,  30,  1.0,  200,  300, 1, "human|warrior",      "4a90d9"],
        [2, "精灵法师",  "common",    1,  25,  0.8,  280,  250, 2, "elf|mage",           "7ed321"],
        [3, "兽人战士",  "common",    1,  40,  0.7,  120,  400, 2, "beast|warrior",      "f5a623"],
        [4, "人类骑士",  "rare",      2,  50,  0.9,  150,  500, 2, "human|warrior",      "9b59b6"],
        [5, "精灵猎手",  "rare",      2,  45,  1.2,  250,  350, 2, "elf|warrior",        "1abc9c"],
        [6, "法师学徒",  "rare",      2,  35,  0.6,  320,  300, 2, "human|mage",         "e74c3c"],
        [7, "野兽猎手",  "epic",      3,  65,  1.5,  180,  450, 2, "beast|warrior",      "e67e22"],
        [8, "精灵祭司",  "epic",      3,  55,  0.5,  300,  350, 2, "elf|mage",           "8e44ad"],
    ]

    ws.append(types)
    ws.append(comments)
    ws.append(fields)
    for i, row in enumerate(data):
        ws.append(row)

    apply_header_style(ws, 1, COLOR_TYPE_ROW)
    apply_header_style(ws, 2, COLOR_COMMENT_ROW)
    apply_header_style(ws, 3, COLOR_FIELD_ROW)
    for i in range(len(data)):
        apply_data_style(ws, i + 4, i % 2 == 1)

    set_col_width(ws, [6, 14, 10, 10, 8, 12, 14, 8, 10, 22, 12])
    freeze_header(ws)
    return wb

# ============================================================
# 表 2: 羁绊配置表 (synergy.xlsx)
# ============================================================
def create_synergy_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "synergy"

    types =    ["int",   "string",   "string",   "int",      "int",      "array_str",       "array_str",        "array_str",         "array_str"]
    comments = ["编号",  "羁绊名称",  "标签键",   "激活阈值1", "激活阈值2", "效果描述(|分隔)", "攻击加成%(|分隔)", "攻速加成%(|分隔)", "范围加成%(|分隔)"]
    fields =   ["id",   "name",      "tag",      "tier1",    "tier2",    "descriptions",    "atk_bonus",        "spd_bonus",         "range_bonus"]

    data = [
        [1, "人类",  "human",    2, 4, "攻击力+10%|攻击力+25%",  "10|25",   "0|0",    "0|0"  ],
        [2, "精灵",  "elf",      2, 4, "攻速+15%|攻速+30%范围+10%", "0|0",  "15|30",  "0|10" ],
        [3, "战士",  "warrior",  2, 4, "攻击+8%攻速+8%|攻击+20%攻速+20%", "8|20", "8|20", "0|0"],
        [4, "法师",  "mage",     2, 4, "范围+20%|范围+50%",    "0|0",     "0|0",    "20|50"],
        [5, "野兽",  "beast",    2, 4, "攻速+20%|攻速+50%",   "0|0",     "20|50",   "0|0"  ],
    ]

    ws.append(types)
    ws.append(comments)
    ws.append(fields)
    for i, row in enumerate(data):
        ws.append(row)

    apply_header_style(ws, 1, COLOR_TYPE_ROW)
    apply_header_style(ws, 2, COLOR_COMMENT_ROW)
    apply_header_style(ws, 3, COLOR_FIELD_ROW)
    for i in range(len(data)):
        apply_data_style(ws, i + 4, i % 2 == 1)

    set_col_width(ws, [6, 10, 10, 10, 10, 32, 16, 16, 16])
    freeze_header(ws)
    return wb

# ============================================================
# 表 3: 波次配置表 (wave.xlsx)
# ============================================================
def create_wave_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "wave"

    types =    ["int",   "int",        "int",     "float",     "float",       "int",      "int"]
    comments = ["波次",  "敌人数量",   "基础HP",  "移速(px/s)", "生成间隔(秒)", "击杀奖励", "波次奖励"]
    fields =   ["id",   "enemy_count","base_hp", "speed",      "spawn_interval","kill_reward","wave_reward"]

    data = []
    for n in range(1, 11):
        enemy_count = 5 + n * 2
        base_hp     = 80 + n * 30
        speed       = 70 + n * 5
        interval    = max(0.5, round(1.5 - n * 0.05, 2))
        kill_reward = 2 + n
        wave_reward = 10 + n * 2
        data.append([n, enemy_count, base_hp, speed, interval, kill_reward, wave_reward])

    ws.append(types)
    ws.append(comments)
    ws.append(fields)
    for i, row in enumerate(data):
        ws.append(row)

    apply_header_style(ws, 1, COLOR_TYPE_ROW)
    apply_header_style(ws, 2, COLOR_COMMENT_ROW)
    apply_header_style(ws, 3, COLOR_FIELD_ROW)
    for i in range(len(data)):
        apply_data_style(ws, i + 4, i % 2 == 1)

    set_col_width(ws, [6, 10, 10, 12, 14, 10, 10])
    freeze_header(ws)
    return wb

# ============================================================
# 表 4: 敌人配置表 (enemy.xlsx)
# ============================================================
def create_enemy_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "enemy"

    types    = ["int",   "string",   "int",   "float",     "int",        "int",      "string"]
    comments = ["编号",  "敌人名称",  "基础HP", "基础移速", "击杀奖励(金)",  "碰触扣血", "颜色(RRGGBB)"]
    fields   = ["id",   "name",      "hp",    "speed",     "kill_reward", "life_damage", "color"]

    data = [
        [1, "普通小兵",   80,  70,  2, 1, "e74c3c"],
        [2, "快速兵",    60,  120, 3, 1, "e67e22"],
        [3, "重甲兵",    200,  50,  5, 2, "95a5a6"],
        [4, "精英兵",    150, 80,  8, 2, "9b59b6"],
        [5, "首领",      500, 60, 20, 5, "2c3e50"],
    ]

    ws.append(types)
    ws.append(comments)
    ws.append(fields)
    for i, row in enumerate(data):
        ws.append(row)

    apply_header_style(ws, 1, COLOR_TYPE_ROW)
    apply_header_style(ws, 2, COLOR_COMMENT_ROW)
    apply_header_style(ws, 3, COLOR_FIELD_ROW)
    for i in range(len(data)):
        apply_data_style(ws, i + 4, i % 2 == 1)

    set_col_width(ws, [6, 12, 10, 10, 14, 10, 14])
    freeze_header(ws)
    return wb

# ============================================================
# 表 5: 商店配置表 (shop.xlsx)
# ============================================================
def create_shop_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "shop"

    types    = ["int",   "int",        "int",        "int",         "array"]
    comments = ["int", "int",        "int",        "int",         "int array (| split)"]
    fields   = ["id",   "refresh_cost","lock_cost",  "free_refresh", "hero_pool"]

    data = [
        [1, 2, 0, 1, "1|2|3|4|5|6|7|8"],
    ]

    ws.append(types)
    ws.append(comments)
    ws.append(fields)
    for i, row in enumerate(data):
        ws.append(row)

    apply_header_style(ws, 1, COLOR_TYPE_ROW)
    apply_header_style(ws, 2, COLOR_COMMENT_ROW)
    apply_header_style(ws, 3, COLOR_FIELD_ROW)
    for i in range(len(data)):
        apply_data_style(ws, i + 4, i % 2 == 1)

    set_col_width(ws, [6, 12, 12, 16, 30])
    freeze_header(ws)
    return wb

# ============================================================
# 主函数
# ============================================================
if __name__ == "__main__":
    os.makedirs("data", exist_ok=True)

    tables = {
        "data/hero.xlsx":    create_hero_table(),
        "data/synergy.xlsx": create_synergy_table(),
        "data/wave.xlsx":    create_wave_table(),
        "data/enemy.xlsx":   create_enemy_table(),
        "data/shop.xlsx":    create_shop_table(),
    }

    for path, wb in tables.items():
        wb.save(path)
        print(f"[OK] Created: {path}")

    print("\nAll tables created!")
    print("Format:")
    print("  Row1: field type (int/float/string/bool/array/array_str/array_int/dict)")
    print("  Row2: Chinese comment (for reading only)")
    print("  Row3: field name (* prefix = skip, empty = skip)")
    print("  Row4+: data rows")
